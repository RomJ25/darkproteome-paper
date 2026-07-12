"""Ingest the two landmark cohort supplements into the Claim Catalog schema.

Reads the REAL published supplementary tables (HCC Camarena/Albà 2024 + ovarian Raja 2025)
and emits one row per published dark-proteome tumor-antigen *claim*, mapped faithfully to
the 19-column contract. Everything the paper does not machine-report becomes `not reported`
(a finding, not a blank).

Depends on openpyxl (ingestion only; the core auditor stays stdlib). Inputs are the
downloaded supplements (see data/SOURCES.md for provenance/URLs).

    python3 src/darkproteome/ingest_cohorts.py [HCC.xlsx] [OVARIAN.xlsx] > /dev/null
"""

import csv
import re
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from schema import COLUMNS, NOT_REPORTED, validate_row  # noqa: E402

import paths  # noqa: E402  -- centralized data paths

NR = NOT_REPORTED
HCC_DEFAULT = paths.HCC_SI
OVA_DEFAULT = paths.RAJA_SI
OUT = os.path.join(paths.REPO, "data", "claim_catalog_real.csv")

HCC_DOI = "10.1126/sciadv.adn3628"
OVA_DOI = "10.1126/sciadv.ads7405"

# Verified from HCC S18 (mouse HHD-DR1 IFN-γ ELISPOT) by direct extraction.
HCC_REACTIVE = {"WMSLDWELYV", "GLFHIYHKI"}          # in-vivo (humanized mouse) reactive
HCC_TCELL_NEG = {"HLWHSATSL", "FLTLQVHGA"}          # assayed, NOT reactive -> validated_negative


# Row-level lineage carried alongside the 19-column contract (underscore = not schema-validated).
# Lets a reader walk any headline numerator back to the exact source cell that produced it.
LINEAGE = ["_canonical", "_claim_unit", "_source_row", "_source_gene", "_pep_index",
           "_n_peps_in_source_row", "_source_class_raw"]


def blank_row():
    r = {c: NR for c in COLUMNS}
    r["meets_consensus_bar_as_reported"] = "insufficient-info"
    r["source_provenance"] = "primary"
    r["extraction_confidence"] = "high"
    for k in LINEAGE:
        r[k] = NR
    r["_claim_unit"] = "peptide-source-claim"   # default; S23 gene-level rows override
    return r


# --- The canonical positive control: an EXPLICITLY CURATED cancer-testis-antigen panel. ---
#
# This replaces a gene-biotype test (`gene_type == protein_coding` -> "control"), which was
# never a curated panel: it admitted PTH, eleven olfactory receptors, seven keratin-associated
# proteins, GNAT1 and other genes that are not cancer-testis antigens in any sense, and 88 of
# its 144 rows carried no peptide at all.
#
# INCLUSION CRITERION: membership in a well-established cancer-testis antigen family, i.e. a
# gene whose expression is restricted to testis (and/or placenta) among normal adult tissues
# and which is re-expressed in tumours. These are the classical CT families used as
# known-real antigen positive controls in the immunopeptidomics literature.
#
# The control answers exactly ONE question: does a KNOWN-REAL antigen also fail the reusable
# reporting bar? It is NOT a positive control for source-attribution resolution.
#
# NOTE: this list is a curation and needs domain sign-off before publication. It is deliberately
# explicit (not a rule) so that a reader can check every member. Genes matched by family prefix.
CTA_FAMILY_PREFIXES = (
    "MAGEA", "MAGEB", "MAGEC",   # MAGE-A/B/C
    "SSX",                       # synovial sarcoma X
    "GAGE",                      # G antigen
    "CT45A",                     # cancer/testis 45
    "CTAG",                      # NY-ESO-1 (CTAG1B) / LAGE-1 (CTAG2)
    "PRAME",                     # PRAME + PRAMEF*
    "PAGE",                      # P antigen
    "SPANX",                     # sperm protein assoc. w/ nucleus, X-linked
    "XAGE",                      # X antigen
    "CSAG",                      # chondrosarcoma-associated
)
CTA_EXACT = {
    "SAGE1",    # sarcoma antigen 1
    "DDX53",    # CAGE
    "TFDP3",    # DP-4
    "ACTL8",    # CT57
    "FTHL17",   # CT38
    "MORC1",    # CT33
    "TDRD1",    # CT41.1
    "CTCFL",    # BORIS / CT27
    "TPTE",     # CT44
}


# A family prefix alone is not enough: MAGEA4-AS1 is the ANTISENSE RNA to MAGEA4, not a
# protein-coding cancer-testis antigen. Non-coding relatives of CT genes are excluded.
CTA_NONCODING_SUFFIXES = ("-AS1", "-AS2", "-AS3", "-OT1", "-IT1", "-DT", "P")


def is_cta_control(gene_name, biotype=None):
    """True iff the gene is a member of the curated cancer-testis-antigen control panel.

    Requires BOTH (a) membership in an established CT family and (b) that the source did not
    label it non-coding -- a CT family's antisense RNA or pseudogene is not a known-real antigen.
    """
    g = (gene_name or "").strip().upper()
    if not g:
        return False
    if any(g.endswith(s) for s in CTA_NONCODING_SUFFIXES if s != "P"):
        return False
    b = (biotype or "").strip().lower()
    if b and ("pseudogene" in b or b in ("lncrna", "lincrna", "antisense")):
        return False
    if g in CTA_EXACT:
        return True
    return any(g.startswith(p) for p in CTA_FAMILY_PREFIXES)


def map_class(biotype, gen_location=None):
    """Map a source-reported biotype to a manuscript ORF class.

    Returns (class, is_coding). `is_coding` says only that the SOURCE labelled the gene
    protein-coding -- it is NOT a claim that the gene is a known-real antigen. Control
    membership is decided separately and explicitly by `is_cta_control`.
    """
    g = (biotype or "").strip().lower()
    loc = (gen_location or "").strip().lower()
    if "pseudogene" in g:
        return "pseudogene-ORF", False
    if g in ("lncrna", "lincrna", "antisense", "bidirectional_promoter_lncrna",
             "sense_intronic", "processed_transcript", "macro_lncrna",
             "3prime_overlapping_ncrna", "sense_overlapping"):
        return "lncRNA-ORF", False
    if g in ("protein_coding", "protein-coding"):
        if loc == "genebody":          # ovarian: cryptic peptide off-frame in a coding gene body
            return "altORF", False
        return "other", True           # source-labelled coding; control status decided separately
    return "other", False


# S23's peptide cells carry PTM annotations inline: "M(Oxidation)PSPIDHPR", "M(+15.99)ESKELVLEL".
# The previous tokenizer required `s.isalpha()`, so ANY peptide bearing a modification was
# silently dropped in full -- e.g. DDX53's only peptide, "PEDLVVM(Oxidation)AE", vanished
# entirely, as did MAGEB3's "LIMKTNM(Oxidation)LVQF". (Splitting on non-letters instead is worse:
# it shatters the peptide AND emits the annotation itself as a fake 9-mer, "OXIDATION".)
# Correct handling: strip the parenthesised annotation, keep the underlying residue.
# A modified and unmodified form of the same peptide collapse to one sequence, which is right --
# they are one peptide sequence, and dedup handles them.
PTM_RE = re.compile(r"\([^)]*\)")


def peptides_from_cell(cell):
    if not cell:
        return []
    out = []
    for tok in PTM_RE.sub("", str(cell)).replace(";", ",").split(","):
        s = tok.strip().upper()
        if s and s.isalpha() and len(s) >= 7:
            out.append(s)
    return out


def build_gene_type_map(wb):
    """gene_name -> gene_type, harvested from HCC S19 (richest gene_type coverage)."""
    m = {}
    if "S19" not in wb.sheetnames:
        return m
    ws = wb["S19"]
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # header
    for r in rows:
        if not r:
            continue
        name, gtype = r[2], r[3]
        if name and gtype and name not in m:
            m[str(name)] = str(gtype)
    return m


def read_s26_qvalues(wb):
    """S26 -- the ONLY table in the corpus reporting per-PSM search statistics.

    Columns: q-value, target_decoy, sp_nc (canonical/noncanonical), Percolator ids, COMET scores.
    The ingester read 4 of this supplement's 33 sheets and never opened it, which is part of why
    `reported_fdr` looked universally absent.

    TWO THINGS, and keeping them apart is the whole point:

    1. The q-value goes in as `psm_qvalue`, NOT as `reported_fdr`. A per-PSM q-value is not a
       protein/ORF-level FDR. Ingesting it as the latter would make the source-translation
       dimension look adjudicable when it is not -- the same unit conflation that put an HLA
       ligand length into a protein-existence rule.

    2. `accepted_decoys_in_class` (D_N) stays EMPTY, and S26 is the proof that this is the
       literature's gap and not ours: the sheet HAS a `target_decoy` column -- the authors held
       the labels -- and every one of its 43 rows is a `target`. Not one accepted decoy is
       published. So the class-specific error rate is not reconstructible even from the most
       statistically complete table in the entire corpus.
    """
    if "S26" not in wb.sheetnames:
        return {}, {"psms": 0, "targets": 0, "decoys": 0}
    it = wb["S26"].iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else "" for c in next(it)]
    idx = {h: i for i, h in enumerate(hdr)}
    if "q-value" not in idx or "sequence" not in idx:
        return {}, {"psms": 0, "targets": 0, "decoys": 0}
    qmap, stats = {}, {"psms": 0, "targets": 0, "decoys": 0}
    for r in it:
        if not r or not r[0]:
            continue
        # STRIP THE PTM ANNOTATION. S26's `sequence` column carries them inline, exactly like
        # S23's peptide cells -- "IM(Oxidation)KAMQSL". Without PTM_RE this reader silently fails
        # to match 8 of its 36 peptides against the catalog, because "PEDLVVM(OXIDATION)AE" is not
        # "PEDLVVMAE". That is the SAME BUG already fixed in `peptides_from_cell`, and it was
        # reintroduced here in a brand-new reader within two commits of fixing it. Caught only by
        # reconciling S26's peptide count against the catalog -- never assume a new parser is
        # immune to a bug you just fixed in an old one.
        seq = PTM_RE.sub("", str(r[idx["sequence"]] or "")).strip().upper()
        q = r[idx["q-value"]]
        td = str(r[idx.get("target_decoy", -1)] or "").strip().lower() if "target_decoy" in idx else ""
        stats["psms"] += 1
        if td == "target":
            stats["targets"] += 1
        elif td == "decoy":
            stats["decoys"] += 1
        if seq and isinstance(q, (int, float)):
            # best (lowest) q-value per peptide sequence
            if seq not in qmap or q < qmap[seq]:
                qmap[seq] = q
    return qmap, stats


def ingest_hcc(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    gtmap = build_gene_type_map(wb)
    qmap, s26 = read_s26_qvalues(wb)
    print(f"  S26 (per-PSM stats): {s26['psms']} PSMs -- {s26['targets']} target, "
          f"{s26['decoys']} decoy; q-values for {len(qmap)} unique peptides")
    if s26["psms"] and not s26["decoys"]:
        print("    -> ZERO accepted decoys published, though the table HAS a target_decoy column.")
        print("       D_N is unrecoverable even here: `accepted_decoys_in_class` stays empty.")
    rows = []
    s26_attached = set()   # reconciled against qmap at the end -- see the check below

    # --- S23: GENE-level rows, EXPANDED to one claim per peptide. ---
    #
    # S23's `Peptide` cell holds MANY peptides (up to 28). The previous code kept only
    # `peps[0]` and emitted one row per gene, then downstream reported the result as
    # "unique peptide-level" -- silently discarding 468 of 715 peptides (65.5%) and turning a
    # gene-level record into a peptide-level claim. Every peptide now becomes its own
    # peptide-source claim, inheriting the gene-level evidence, with full row lineage.
    #
    # A gene row with NO parseable peptide stays as ONE gene-level record (it is a real claim
    # with no peptide reported); it is tagged `_claim_unit=gene-level` so it can be excluded
    # from peptide-level denominators rather than silently inflating them.
    #
    # `tumor_specificity_basis` was previously hardcoded to "broad-normal-panel" for every S23
    # row. S23 HAS NO TUMOR-SPECIFICITY COLUMN (header: gene_id, transcript_id, gene_name,
    # gene_type, n, % patients, Source, Peptide, '>10% & highly expressed'). The one flag it
    # carries is a tumour-side frequency flag, not a normal-tissue result. It is now
    # `not reported`, which is what the source record actually supports.
    ws = wb["S23"]
    it = ws.iter_rows(values_only=True)
    next(it, None)  # header
    for src_row, r in enumerate(it, start=2):   # start=2: row 1 is the header
        if not r or not r[0]:
            continue
        gene_id, _tx, gene_name, gene_type, n, _pct, source, pepcell, _hi = (list(r) + [None] * 9)[:9]
        peps = peptides_from_cell(pepcell)
        cls, _coding = map_class(gene_type)
        gname = str(gene_name or gene_id or NR)
        is_ctrl = is_cta_control(gene_name, gene_type)

        def _s23(pep, idx, unit):
            row = blank_row()
            row["peptide_sequence"] = pep
            row["orf_id_or_locus"] = gname
            row["orf_class"] = cls
            row["antigen_type"] = "TAA" if is_ctrl else "TSA"
            row["cancer_type"] = "Hepatocellular Carcinoma"
            row["evidence_types"] = "immunopeptidomics"
            row["n_unique_peptides"] = str(len(peps)) if peps else NR
            # the length of THIS ligand -- not the min over the gene's whole peptide set
            row["ligand_len"] = str(len(pep)) if pep != NR else NR
            row["source_pep_len"] = NR      # S23 reports no source-protein tryptic support
            row["psm_qvalue"] = str(qmap[pep]) if pep in qmap else NR   # from S26; NOT an FDR
            row["tumor_specificity_modality"] = NR   # S23 has no specificity column at all
            row["tumor_specificity_scope"] = NR
            row["tumor_specificity_result"] = NR
            row["validation_level"] = "MS-presented"
            row["citation_doi_pmid"] = HCC_DOI
            row["citation_location"] = "Table S23"
            row["_canonical"] = "yes" if is_ctrl else "no"
            row["_claim_unit"] = unit
            row["_source_row"] = str(src_row)
            row["_source_gene"] = gname
            row["_pep_index"] = str(idx)
            row["_n_peps_in_source_row"] = str(len(peps))
            row["_source_class_raw"] = str(gene_type or NR)
            return row

        if peps:
            for i, p in enumerate(peps):
                rows.append(_s23(p, i, "peptide-source-claim"))
        else:
            rows.append(_s23(NR, -1, "gene-level"))

    # --- S16/S17/S18: the 13 experimentally-tested peptides (immunogenicity seed) ---
    ws = wb["S16"]
    it = ws.iter_rows(values_only=True)
    next(it, None)  # header
    for r in it:
        if not r or not r[0]:
            continue
        pep, _gid, gene_name, _orf, _pos, _eid, _rank, _aff, binding, tumorspec, _np, riborf = (list(r) + [None] * 12)[:12]
        pep = str(pep).strip().upper()
        cls, canonical = map_class(gtmap.get(str(gene_name), ""))
        if pep in HCC_REACTIVE:
            val = "in-vivo"
        elif pep in HCC_TCELL_NEG:
            val = "validated_negative"
        else:
            val = "MS-presented"
        ev = "immunopeptidomics, MHC binding assay, ELISPOT"
        if str(riborf).strip().upper() == "YES":
            ev = "Ribo-seq, " + ev
        row = blank_row()
        row["peptide_sequence"] = pep
        row["orf_id_or_locus"] = str(gene_name or NR)
        row["orf_class"] = cls
        row["antigen_type"] = "TSA"
        row["cancer_type"] = "Hepatocellular Carcinoma"
        row["hla_allele"] = "HLA-A*02:01"
        row["evidence_types"] = ev
        row["ligand_len"] = str(len(pep))
        row["source_pep_len"] = NR
        row["psm_qvalue"] = str(qmap[pep]) if pep in qmap else NR   # from S26; NOT an FDR
        # S16 reports a per-peptide `Tumor-specific` Yes/No, so the SCOPE is a genuine per-claim
        # result. The MODALITY is not stated in S16 -- but S19 is the table that defines it:
        # its columns are `Normal adjacent RNA-Seq Expression (FPKM)`, `Tumor RNA-Seq Expression
        # (FPKM)` and `tumor-specificity`. So the paper's tumour-specific call is an RNA-Seq
        # comparison against the patient's own ADJACENT normal tissue.
        #
        # That is RNA (not presentation) against ONE matched organ (not a broad panel), so it
        # cannot support "absent from normal HLA presentation". It scores weak-pass. The original
        # ingester hardcoded it to `broad-normal-panel` and strict-passed it, which was wrong on
        # both axes at once. Not a guess: read off S19's own column headers.
        ts = str(tumorspec).strip().lower()
        # Keep the NOs. A reported "not tumour-specific" is an explicit normal DETECTION -- the
        # only true empirical negative in the corpus. The old code recorded only the YESes and
        # dropped the NOs to `not reported`, which made a record that had SPOKEN look SILENT.
        if ts in ("yes", "no"):
            row["tumor_specificity_modality"] = "normal-rna-matched"
            row["tumor_specificity_scope"] = "per-claim-reported"
            row["tumor_specificity_result"] = ("absent-from-normal" if ts == "yes"
                                               else "detected-in-normal")
        else:
            row["tumor_specificity_modality"] = NR
            row["tumor_specificity_scope"] = NR
            row["tumor_specificity_result"] = NR
        row["validation_level"] = val
        row["citation_doi_pmid"] = HCC_DOI
        row["citation_location"] = "Tables S16/S17/S18"
        row["_canonical"] = "no"
        rows.append(row)

    wb.close()

    # RECONCILE THE PARSER AGAINST THE CATALOG. A parser that silently fails to match is exactly
    # how the PTM bug survived: `read_s26_qvalues` at first did not strip "(Oxidation)", so 8 of
    # its 36 peptides could never match a catalog peptide, and nothing complained -- the q-values
    # simply never attached. This makes that condition loud. Any unattached q-value is either a
    # parse bug on our side or an S26 peptide genuinely absent from the claim tables; either way
    # it must be SEEN, not assumed away.
    attached = {r["peptide_sequence"] for r in rows if r.get("psm_qvalue") not in (NR, None, "")}
    unattached = sorted(set(qmap) - attached)
    print(f"  S26 reconciliation: {len(attached)}/{len(qmap)} q-value peptides attached to claims")
    if unattached:
        print(f"    !! {len(unattached)} S26 peptides matched NO claim: {unattached[:6]}"
              f"{' ...' if len(unattached) > 6 else ''}")
        print( "       If these look like PTM-annotated forms, the reader is not stripping them.")
    return rows


def _find_header(ws, needles):
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(c).strip().lower() if c is not None else "" for c in r]
        if all(any(n in v for v in vals) for n in needles):
            return i, vals
    return None, None


def ingest_ovarian(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = []
    pep_biotype = {}

    # --- S2: all identified cryptic peptides (presentation claims) ---
    ws = wb["Supplementary Table S2"]
    all_rows = list(ws.iter_rows(values_only=True))
    hdr = all_rows[0]
    cols = {str(c).strip(): j for j, c in enumerate(hdr) if c}
    ci_pep = cols.get("Peptide")
    ci_bio = cols.get("Biotype")
    ci_loc = cols.get("Gen_Location")
    ci_gene = cols.get("Gene Symbol")
    s2_peps = []
    for r in all_rows[1:]:
        if not r or ci_pep is None or not r[ci_pep]:
            continue
        pep = str(r[ci_pep]).strip().upper()
        biotype = r[ci_bio] if ci_bio is not None else None
        genloc = r[ci_loc] if ci_loc is not None else None
        gene = r[ci_gene] if ci_gene is not None else None
        pep_biotype[pep] = (biotype, genloc)
        s2_peps.append((pep, biotype, genloc, gene))

    # --- S3: the 38 prioritized + immunogenicity-tested candidates ---
    ws3 = wb["Supplementary Table S3"]
    hi, hvals = _find_header(ws3, ["sequence", "predicted hla", "gene symbol"])
    s3_set = set()
    if hi is not None:
        idx = {name: j for j, name in enumerate(hvals)}
        def col(part):
            for name, j in idx.items():
                if part in name:
                    return j
            return None
        j_seq, j_gene, j_hla = col("sequence"), col("gene symbol"), col("predicted hla")
        for r in list(ws3.iter_rows(values_only=True))[hi + 1:]:
            if not r or j_seq is None or not r[j_seq]:
                continue
            pep = str(r[j_seq]).strip().upper()
            if not pep.isalpha():
                continue
            s3_set.add(pep)
            biotype, genloc = pep_biotype.get(pep, (None, None))
            cls, _ = map_class(biotype, genloc)
            row = blank_row()
            row["peptide_sequence"] = pep
            row["orf_id_or_locus"] = str(r[j_gene]) if j_gene is not None and r[j_gene] else NR
            row["orf_class"] = cls
            row["antigen_type"] = "TSA"
            row["cancer_type"] = "Metastatic Ovarian Cancer"
            row["hla_allele"] = str(r[j_hla]) if j_hla is not None and r[j_hla] else NR
            row["evidence_types"] = "immunopeptidomics, T-cell assay (result figure-locked)"
            row["ligand_len"] = str(len(pep))
            row["source_pep_len"] = NR
            # The authors PRIORITIZED this table using HLA Ligand Atlas -- a BROAD normal
            # ligandome, so the measurement and the panel are both right. What is missing is the
            # scope: it is an INCLUSION CRITERION for the table, not a per-peptide reported
            # result. Every row carries it unconditionally and no per-peptide value can be
            # re-derived or contested. Weak-pass, not strict -- which is the whole reason `scope`
            # exists as a field. This is the closest any claim in the corpus comes to strict.
            row["tumor_specificity_modality"] = "normal-ligandome-broad"
            row["tumor_specificity_scope"] = "cohort-inclusion-criterion"
            row["tumor_specificity_result"] = "absent-from-normal"   # implied by inclusion
            row["validation_level"] = NR  # T-cell tested, but per-peptide reactive/non-reactive only in Fig 6
            row["citation_doi_pmid"] = OVA_DOI
            row["citation_location"] = "Table S3 + Fig 6"
            row["_canonical"] = "no"
            rows.append(row)

    # S2 claims, excluding the 38 already added at higher resolution from S3
    for pep, biotype, genloc, gene in s2_peps:
        if pep in s3_set:
            continue
        cls, _ = map_class(biotype, genloc)
        row = blank_row()
        row["peptide_sequence"] = pep
        row["orf_id_or_locus"] = str(gene) if gene else NR
        row["orf_class"] = cls
        row["antigen_type"] = "TSA"
        row["cancer_type"] = "Metastatic Ovarian Cancer"
        row["evidence_types"] = "immunopeptidomics"
        row["ligand_len"] = str(len(pep))
        row["source_pep_len"] = NR
        row["tumor_specificity_modality"] = NR   # full identified set, not specificity-filtered
        row["tumor_specificity_scope"] = NR
        row["tumor_specificity_result"] = NR
        row["validation_level"] = "MS-presented"
        row["citation_doi_pmid"] = OVA_DOI
        row["citation_location"] = "Table S2"
        row["_canonical"] = "no"
        rows.append(row)

    wb.close()
    return rows


def main():
    hcc_path = sys.argv[1] if len(sys.argv) > 1 else HCC_DEFAULT
    ova_path = sys.argv[2] if len(sys.argv) > 2 else OVA_DEFAULT
    paths.require(hcc_path, ova_path)
    rows = ingest_hcc(hcc_path) + ingest_ovarian(ova_path)

    bad = 0
    for r in rows:
        probs = validate_row({k: v for k, v in r.items() if not k.startswith("_")})
        if probs:
            bad += 1
            if bad <= 5:
                print(f"  !! invalid: {r.get('peptide_sequence')}: {probs}", file=sys.stderr)

    # --- UNIT INVARIANTS. The peps[0] bug was silent; these make it loud. ---
    s23 = [r for r in rows if r["citation_location"] == "Table S23"]
    pep_claims = [r for r in s23 if r["_claim_unit"] == "peptide-source-claim"]
    gene_only = [r for r in s23 if r["_claim_unit"] == "gene-level"]
    src_rows = {r["_source_row"] for r in s23}
    # every peptide in a multi-peptide source cell must survive into its own claim
    expected = sum(int(r["_n_peps_in_source_row"]) for r in s23 if r["_pep_index"] == "0")
    assert len(pep_claims) == expected, \
        f"peptide expansion lost rows: {len(pep_claims)} claims vs {expected} source peptides"
    # a gene-level record must never carry a peptide (it would inflate peptide denominators)
    assert all(r["peptide_sequence"] == NR for r in gene_only), "gene-level row carries a peptide"
    # every numerator must be a subset of its denominator: no claim without lineage
    assert all(r["_source_row"] != NR for r in s23), "S23 row missing lineage"

    fieldnames = COLUMNS + LINEAGE
    with open(OUT, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, NR) for k in fieldnames})

    from collections import Counter
    print(f"wrote {len(rows)} claims to {OUT}  (schema-invalid: {bad})")
    print("by cohort:", dict(Counter(r["cancer_type"] for r in rows)))
    print("by orf_class:", dict(Counter(r["orf_class"] for r in rows)))
    print("by validation_level:", dict(Counter(r["validation_level"] for r in rows)))
    print()
    print("--- HCC S23 expansion (was: 1 row per gene, first peptide only) ---")
    print(f"  source gene rows            : {len(src_rows)}")
    print(f"  peptide-source claims       : {len(pep_claims)}")
    print(f"  gene-level rows (no peptide): {len(gene_only)}")
    print(f"  unique peptide sequences    : {len({r['peptide_sequence'] for r in pep_claims})}")
    print(f"  duplicate peptide-source claims: "
          f"{len(pep_claims) - len({(r['_source_gene'], r['peptide_sequence']) for r in pep_claims})}")
    print()
    ctrl = [r for r in rows if r["_canonical"] == "yes"]
    print("--- canonical CTA control (curated panel, was: every protein_coding row) ---")
    print(f"  control claims: {len(ctrl)}  across {len({r['_source_gene'] for r in ctrl})} genes")
    print(f"  genes: {sorted({r['_source_gene'] for r in ctrl})}")


if __name__ == "__main__":
    main()
